import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _h(cell, bg="FF1565C0", fg="FFFFFFFF"):
    cell.font = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="FFBDBDBD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _b(cell):
    thin = Side(style="thin", color="FFBDBDBD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def generate(incomes, expenses, loans, period_label="Текущий период"):
    wb = Workbook()

    # ── Сводка ────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Сводка"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    ws.merge_cells("A1:B1")
    ws["A1"] = f"Финансовый отчёт — {period_label}"
    ws["A1"].font = Font(bold=True, size=13, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")

    total_i = sum(x["amount"] for x in incomes)
    total_e = sum(x["amount"] for x in expenses)
    bal = total_i - total_e

    for row, (label, val, bg) in enumerate([
        ("Доходы", total_i, "FFE8F5E9"),
        ("Расходы", total_e, "FFFFEBEE"),
        ("Баланс", bal, "FFE3F2FD"),
    ], start=3):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = val
        ws[f"B{row}"].number_format = "#,##0.00"
        ws[f"A{row}"].font = Font(bold=True, name="Arial")
        for col in "AB":
            ws[f"{col}{row}"].fill = PatternFill("solid", start_color=bg)
            _b(ws[f"{col}{row}"])

    # Расходы по категориям
    row = 8
    ws[f"A{row}"] = "Категория расходов"; _h(ws[f"A{row}"], "FFC62828")
    ws[f"B{row}"] = "Сумма"; _h(ws[f"B{row}"], "FFC62828")
    cats = {}
    for e in expenses:
        cats[e["category"]] = cats.get(e["category"], 0) + e["amount"]
    for cat, amt in sorted(cats.items(), key=lambda x: -x[1]):
        row += 1
        ws[f"A{row}"] = cat
        ws[f"B{row}"] = amt
        ws[f"B{row}"].number_format = "#,##0.00"
        for col in "AB":
            ws[f"{col}{row}"].fill = PatternFill("solid", start_color="FFFFEBEE")
            _b(ws[f"{col}{row}"])

    # ── Доходы ────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Доходы")
    for col, (h, w) in enumerate(zip(["Дата","Категория","Описание","Сумма"], [14,22,30,16]), 1):
        _h(ws2.cell(1, col, h), "FF2E7D32")
        ws2.column_dimensions[get_column_letter(col)].width = w
    for row, inc in enumerate(incomes, 2):
        bg = "FFE8F5E9" if row % 2 == 0 else "FFFFFFFF"
        for col, val in enumerate([inc.get("date",""), inc["category"], inc.get("desc",""), inc["amount"]], 1):
            c = ws2.cell(row, col, val)
            c.fill = PatternFill("solid", start_color=bg)
            _b(c)
        ws2.cell(row, 4).number_format = "#,##0.00"

    # ── Расходы ───────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Расходы")
    for col, (h, w) in enumerate(zip(["Дата","Категория","Описание","Сумма"], [14,22,30,16]), 1):
        _h(ws3.cell(1, col, h), "FFC62828")
        ws3.column_dimensions[get_column_letter(col)].width = w
    for row, exp in enumerate(expenses, 2):
        bg = "FFFFEBEE" if row % 2 == 0 else "FFFFFFFF"
        for col, val in enumerate([exp.get("date",""), exp["category"], exp.get("desc",""), exp["amount"]], 1):
            c = ws3.cell(row, col, val)
            c.fill = PatternFill("solid", start_color=bg)
            _b(c)
        ws3.cell(row, 4).number_format = "#,##0.00"

    # ── Кредиты ───────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Кредиты")
    hdrs = ["Банк","Название","Сумма","Ставка%","Срок","Ежемес.","Выдан","Статус"]
    widths = [20,20,16,10,8,16,14,12]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        _h(ws4.cell(1, col, h))
        ws4.column_dimensions[get_column_letter(col)].width = w
    for row, loan in enumerate(loans, 2):
        from loan_calc import annuity
        monthly = annuity(loan["amount"], loan["rate"], loan["months"])
        vals = [loan.get("bank",""), loan["name"], loan["amount"], loan["rate"],
                loan["months"], round(monthly,0), loan.get("start_date",""), "Активный" if loan.get("active", True) else "Закрыт"]
        bg = "FFE3F2FD" if row % 2 == 0 else "FFFFFFFF"
        for col, val in enumerate(vals, 1):
            c = ws4.cell(row, col, val)
            c.fill = PatternFill("solid", start_color=bg)
            _b(c)
        ws4.cell(row, 3).number_format = "#,##0.00"
        ws4.cell(row, 6).number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
